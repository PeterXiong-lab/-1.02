import os
import sys
import pandas as pd
from openpyxl import load_workbook

MUNICIPALITIES = {'北京市', '上海市', '天津市', '重庆市'}


# ======================================================
# 地址解析
# ======================================================

def parse_address(addr):

    addr = str(addr).strip()

    if addr[:3] in MUNICIPALITIES:
        province = addr[:3]
        pe = 3

    elif '自治区' in addr:
        pe = addr.index('自治区') + 3
        province = addr[:pe]

    elif '省' in addr:
        pe = addr.index('省') + 1
        province = addr[:pe]

    else:
        province = addr[:3]
        pe = 3

    try:
        ce = addr.index('市', pe) + 1
    except:
        ce = pe

    city = addr[pe:ce]

    if not city:
        city = province

    endings = ['区', '县', '镇', '乡', '旗']

    de = len(addr)

    for ch in endings:

        try:
            pos = addr.index(ch, ce) + 1

            if pos < de:
                de = pos

        except:
            pass

    try:
        pos = addr.index('市', ce) + 1

        if pos < de:
            de = pos

    except:
        pass

    district = addr[ce:de]
    detail = addr[de:]

    for prefix in [addr[:de], city + district, district]:

        if prefix and detail.startswith(prefix):
            detail = detail[len(prefix):]
            break

    return province, city, district, detail


# ======================================================
# 自动寻找文件
# ======================================================

CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))

# 12
path_12 = None

if os.path.exists(os.path.join(CURRENT_DIR, "12.xlsx")):
    path_12 = os.path.join(CURRENT_DIR, "12.xlsx")

# 123
path_123 = None

if os.path.exists(os.path.join(CURRENT_DIR, "123.xlsx")):
    path_123 = os.path.join(CURRENT_DIR, "123.xlsx")

# 检查
if not path_12 or not path_123:

    print("\n【错误】缺少xlsx文件！")
    print("请确保：")
    print("12.xlsx")
    print("123.xlsx")
    print("在同一个文件夹内")

    input("\n按回车退出...")
    sys.exit()

output_path = os.path.join(CURRENT_DIR, "123_output.xlsx")


# ======================================================
# 读取12
# ======================================================

print("=================== 第一步：读取12数据 ===================")

df_12 = pd.read_excel(path_12, dtype=str)

df_12.columns = df_12.columns.str.strip()

# 全部转字符串，防止SKU/手机号/单号精度丢失
df_12 = df_12.fillna('')

print("12表字段：")
print(list(df_12.columns))


# ======================================================
# 打开123模板（保留公式）
# ======================================================

print("\n=================== 第二步：打开123模板 ===================")

wb = load_workbook(path_123)

ws = wb.active

# 获取表头

headers_123 = {}

for col in range(1, ws.max_column + 1):

    header = ws.cell(1, col).value

    if header:
        headers_123[str(header).strip()] = col

print("123表字段：")
print(list(headers_123.keys()))


# ======================================================
# 行数校验
# ======================================================

excel_data_rows = ws.max_row - 1

if len(df_12) != excel_data_rows:

    print("\n【严重警告】两个表行数不一致！")
    print(f"12表行数：{len(df_12)}")
    print(f"123表数据行数：{excel_data_rows}")

    input("\n请检查后按回车退出...")
    sys.exit()


# ======================================================
# 自动复制字段
# ======================================================

print("\n=================== 第三步：复制字段 ===================")

copy_map = {

    # 基础信息
    '其它出库业务单号': '其它出库业务单号',
    '收货人': '收货人',
    '收货电话': '收货电话',
    '收货地址': '收货地址',

    # 业务字段
    'SKU采购总⾦额（含税）': '单价',
    '采购数量（采购单位）': '数量',
    'SKU编码': 'SKU编码',
}

for source_col, target_col in copy_map.items():

    # 检查12字段
    if source_col not in df_12.columns:

        print(f"\n【警告】12缺少字段：{source_col}")
        continue

    # 检查123字段
    if target_col not in headers_123:

        print(f"\n【警告】123缺少字段：{target_col}")
        continue

    target_excel_col = headers_123[target_col]

    # 写入
    for i in range(len(df_12)):

        excel_row = i + 2

        value = df_12.iloc[i][source_col]

        # ==================================================
        # 编码类字段强制字符串
        # ==================================================

        if target_col in [
            'SKU编码',
            '其它出库业务单号',
            '收货电话'
        ]:

            value = str(value).strip()

            # 去掉 pandas 自动补的 .0
            if value.endswith('.0'):
                value = value[:-2]

        # 写入
        ws.cell(excel_row, target_excel_col).value = value

    print(f"【成功】{source_col} -> {target_col}")


# ======================================================
# 地址解析
# ======================================================

print("\n=================== 第四步：地址解析 ===================")

if '收货地址' in df_12.columns:

    parsed_data = []

    for addr in df_12['收货地址']:

        if str(addr).strip():

            parsed_data.append(parse_address(addr))

        else:

            parsed_data.append(("", "", "", ""))

    parsed_df = pd.DataFrame(

        parsed_data,

        columns=[
            '收货省份',
            '收货城市',
            '收货区县',
            '收货详细地址'
        ]
    )

    # ==================================================
    # 备注
    # ==================================================

    clean_addresses = (

        df_12['收货地址']
        .astype(str)
        .fillna('')
        .str.replace('nan', '')
        .str.strip()
    )

    remark_series = (

        df_12['收货人'].astype(str).fillna('').str.replace('nan', '')
        + " "
        + df_12['收货电话'].astype(str).fillna('').str.replace('nan', '')
        + " "
        + clean_addresses

    ).str.strip()

    parsed_df['备注'] = remark_series

    # ==================================================
    # 自动新增列并写入
    # ==================================================

    for field in parsed_df.columns:

        # 不存在则自动新增
        if field not in headers_123:

            new_col = ws.max_column + 1

            ws.cell(1, new_col).value = field

            headers_123[field] = new_col

            print(f"【自动新增列】{field}")

        target_excel_col = headers_123[field]

        # 写入数据
        for i in range(len(parsed_df)):

            excel_row = i + 2

            value = parsed_df.iloc[i][field]

            ws.cell(excel_row, target_excel_col).value = value

        print(f"【成功】写入：{field}")


# ======================================================
# 保存
# ======================================================

print("\n=================== 第五步：保存文件 ===================")

wb.save(output_path)

print("\n【✨ 全部完成 ✨】")
print("公式 / XLOOKUP / 格式 已完整保留")
print("SKU编码精度已保留")
print("地址解析 / 自动新增列 已完成")
print(f"输出文件：{output_path}")

input("\n按回车退出...")
